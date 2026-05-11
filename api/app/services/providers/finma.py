"""
Raven — FINMA Supervised Institutions Provider

Downloads and parses the official FINMA Excel list of all
banks and securities firms authorised in Switzerland.

Source: https://www.finma.ch/en/finma-public/authorised-institutions-individuals-and-products/
Excel: https://www.finma.ch/en/~/media/finma/dokumente/bewilligungstraeger/xlsx/beh.xlsx
PDF:   https://www.finma.ch/en/~/media/finma/dokumente/bewilligungstraeger/pdf/beh.pdf

Updated: monthly by FINMA
No auth required. Public data.
"""

import httpx
import io
import re
from datetime import datetime
from typing import Optional

FINMA_XLSX_URL = "https://www.finma.ch/en/~/media/finma/dokumente/bewilligungstraeger/xlsx/beh.xlsx"
FINMA_PDF_URL  = "https://www.finma.ch/en/~/media/finma/dokumente/bewilligungstraeger/pdf/beh.pdf"
FINMA_PAGE_URL = "https://www.finma.ch/en/finma-public/authorised-institutions-individuals-and-products/"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept":     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/octet-stream, */*",
    "Referer":    FINMA_PAGE_URL,
}

# Cache the parsed institution list (refresh daily)
_CACHE: dict = {}
_CACHE_TTL = 3600 * 24  # 24h


def _download_xlsx() -> Optional[bytes]:
    """Download FINMA Excel file."""
    try:
        r = httpx.get(FINMA_XLSX_URL, headers=HEADERS, timeout=20, follow_redirects=True)
        if r.status_code == 200:
            return r.content
        print(f"[finma] Excel download: HTTP {r.status_code}")
    except Exception as e:
        print(f"[finma] Excel download error: {e}")
    return None


def _parse_xlsx(data: bytes) -> list:
    """
    Parse FINMA Excel into list of institution dicts.
    FINMA Excel columns (beh.xlsx) typically include:
    - Institution name (German/French/Italian)
    - Licence type
    - Location/domicile
    - Authorisation date
    - Status (active/withdrawn)
    - FINMA profile URL
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

        institutions = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row — look for row containing name-like column
            header_row = None
            header_idx = 0
            for i, row in enumerate(rows[:10]):
                row_str = [str(c or "").lower() for c in row]
                if any(k in " ".join(row_str) for k in ("name", "firma", "raison", "institution", "bank")):
                    header_row = row
                    header_idx = i
                    break

            if header_row is None:
                continue

            # Map column indices
            headers = [str(h or "").strip().lower() for h in header_row]

            def col(keywords):
                for kw in keywords:
                    for i, h in enumerate(headers):
                        if kw in h:
                            return i
                return None

            name_col    = col(["firma", "name", "institution", "raison"])
            type_col    = col(["kategorie", "type", "bewilligung", "licence", "kategori"])
            status_col  = col(["status", "zustand", "état"])
            city_col    = col(["ort", "city", "lieu", "domicile", "sitz"])
            date_col    = col(["datum", "date", "seit", "since", "autorisation"])

            if name_col is None:
                continue

            for row in rows[header_idx + 1:]:
                if not row or not row[name_col]:
                    continue
                name_val = str(row[name_col] or "").strip()
                if not name_val or len(name_val) < 2:
                    continue

                inst = {
                    "name":          name_val,
                    "licence_type":  str(row[type_col] or "").strip() if type_col is not None else "",
                    "status":        str(row[status_col] or "").strip() if status_col is not None else "",
                    "city":          str(row[city_col] or "").strip() if city_col is not None else "",
                    "auth_date":     str(row[date_col] or "").strip() if date_col is not None else "",
                    "sheet":         sheet_name,
                }
                institutions.append(inst)

        return institutions
    except ImportError:
        print("[finma] openpyxl not installed")
    except Exception as e:
        print(f"[finma] Excel parse error: {e}")
    return []


def _get_institutions() -> list:
    """Get cached or fresh FINMA institution list."""
    if "institutions" in _CACHE:
        entry = _CACHE["institutions"]
        if (datetime.utcnow() - entry["ts"]).seconds < _CACHE_TTL:
            return entry["data"]

    # Non-blocking: if download takes too long, return empty and retry later
    import threading
    if _CACHE.get("_downloading"):
        return _CACHE.get("institutions", {}).get("data", [])

    _CACHE["_downloading"] = True
    try:
        data = _download_xlsx()
        if data:
            institutions = _parse_xlsx(data)
            # Free the raw bytes immediately after parsing
            del data
            if institutions:
                _CACHE["institutions"] = {"data": institutions, "ts": datetime.utcnow()}
                print(f"[finma] Loaded {len(institutions)} institutions from Excel")
                _CACHE["_downloading"] = False
                return institutions
    except Exception as e:
        print(f"[finma] Institution load error: {e}")
    finally:
        _CACHE["_downloading"] = False

    return _CACHE.get("institutions", {}).get("data", [])


def preload_async():
    """Preload FINMA Excel in a background thread at startup."""
    import threading
    t = threading.Thread(target=_get_institutions, daemon=True)
    t.start()


def _find_institution(display_name: str, slug: str = "") -> Optional[dict]:
    """Search FINMA institution list by name."""
    institutions = _get_institutions()
    if not institutions:
        return None

    name_lower = display_name.lower().strip()
    slug_lower = slug.lower().strip()

    # Exact match first
    for inst in institutions:
        if inst["name"].lower() == name_lower:
            return inst

    # Substring match
    best = None
    best_score = 0
    for inst in institutions:
        inst_lower = inst["name"].lower()
        # Check both directions
        if name_lower in inst_lower:
            score = len(name_lower) / len(inst_lower)
        elif inst_lower in name_lower:
            score = len(inst_lower) / len(name_lower)
        else:
            # Check slug
            score = 0
            if slug_lower and len(slug_lower) >= 4:
                slug_parts = slug_lower.replace("-", " ").split()
                matches = sum(1 for part in slug_parts if part in inst_lower)
                score = matches / max(len(slug_parts), 1) * 0.6

        if score > best_score and score >= 0.5:
            best_score = score
            best = inst

    return best


def enrich_counterparty(slug: str, display_name: str = "") -> dict:
    """
    Main entry point. Look up entity in FINMA Excel register.
    Returns licence type, status, city, authorisation date.
    """
    result = {
        "source":     "finma_xlsx",
        "available":  False,
        "fetched_at": datetime.utcnow().isoformat(),
    }

    inst = _find_institution(display_name or slug, slug)

    if not inst:
        # Try with just first significant word
        first_word = (display_name or slug).split()[0] if (display_name or slug) else ""
        if len(first_word) >= 4:
            inst = _find_institution(first_word)

    if not inst:
        result["reason"] = "not_found_in_finma_xlsx"
        return result

    # Determine if licence is active
    status_lower = inst.get("status", "").lower()
    active_terms = ("aktiv", "active", "bewilligt", "authorised", "autorisé", "autorizzato", "")
    withdrawn    = ("entzogen", "withdrawn", "widerrufen", "révoqué", "revocato", "expired")
    is_active    = not any(t in status_lower for t in withdrawn)

    result.update({
        "available":         True,
        "license_active":    is_active,
        "finma_legal_name":  inst["name"],
        "finma_licence_type": inst.get("licence_type") or inst.get("sheet", ""),
        "finma_status":      inst.get("status", ""),
        "finma_city":        inst.get("city", ""),
        "finma_auth_date":   inst.get("auth_date", ""),
        "finma_source":      "beh.xlsx",
        "finma_url":         FINMA_PAGE_URL,
        "xlsx_url":          FINMA_XLSX_URL,
    })

    if inst.get("auth_date"):
        # Try to compute years regulated
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y", "%d/%m/%Y"):
            try:
                d = datetime.strptime(inst["auth_date"][:10], fmt)
                result["years_regulated"] = (datetime.utcnow() - d).days // 365
                break
            except Exception:
                continue

    return result
